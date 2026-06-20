import openpyxl

EXCEL_PATH = "/Users/tirthpatel/Downloads/Report(3).xlsx"

def main():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        print("Sheet names:", wb.sheetnames)
        
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"\nSheet '{name}' dimensions: {sheet.dimensions}")
            
            # Print headers (row 1)
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
                
            print("Headers (first 30 cols):", rows[0][:30])
            
            # Print first 5 data rows
            print("First 5 data rows:")
            for idx, r in enumerate(rows[1:6], 2):
                print(f"Row {idx}: {r[:10]} ...")
                
            # Search for Mishwa's email or name in all rows
            print("\nSearching for Mishwa's email or name...")
            found = False
            for idx, r in enumerate(rows, 1):
                r_str = str(r).lower()
                if "mishwa" in r_str or "hansaliya" in r_str or "hansiya" in r_str or "141744" in r_str:
                    print(f"Row {idx}: {r}")
                    found = True
            if not found:
                print("No matches found for Mishwa in this sheet.")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()
