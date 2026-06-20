import os
import openpyxl

DOWNLOADS_DIR = "/Users/tirthpatel/Downloads"

def search_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for col_idx, cell_value in enumerate(row, 1):
                    if cell_value:
                        val_str = str(cell_value).lower()
                        if "mishwa" in val_str or "hansaliya" in val_str or "hansiya" in val_str or "141744" in val_str:
                            print(f"Match found in {file_path} | Sheet: {sheet_name} | Row {row_idx}, Col {col_idx} | Value: {cell_value}")
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

def main():
    print("Searching recipient_certificates Excel files in Downloads...")
    for file in os.listdir(DOWNLOADS_DIR):
        if "recipient_certificates_report" in file and file.endswith(".xlsx"):
            full_path = os.path.join(DOWNLOADS_DIR, file)
            search_excel(full_path)
    print("Recipient Excel scan complete.")

if __name__ == "__main__":
    main()
