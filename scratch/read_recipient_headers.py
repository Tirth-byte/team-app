import openpyxl

EXCEL_PATH = "/Users/tirthpatel/Downloads/recipient_certificates_report.xlsx"

def main():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        print("Sheets:", wb.sheetnames)
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"Sheet: {name} | Dimensions: {sheet.dimensions}")
            rows = list(sheet.iter_rows(values_only=True))
            print(f"Total rows: {len(rows)}")
            for idx, r in enumerate(rows[:5]):
                print(f"Row {idx+1}: {r}")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()
