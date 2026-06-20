import os
import openpyxl
import re
from google.cloud import firestore
import firebase_admin
from firebase_admin import credentials, firestore

# Load environment variables
# Load private key from .env.local
env_vars = {}
with open(".env.local", "r") as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env_vars[k.strip()] = v.strip().strip('"').strip("'")

project_id = env_vars.get("FIREBASE_PROJECT_ID")
client_email = env_vars.get("FIREBASE_CLIENT_EMAIL")
private_key = env_vars.get("FIREBASE_PRIVATE_KEY").replace("\\n", "\n")

if not firebase_admin._apps:
    cred = credentials.Certificate({
        "project_id": project_id,
        "client_email": client_email,
        "private_key": private_key
    })
    firebase_admin.initialize_app(cred)

db = firestore.client()
mishwa_uid = "A1gwf0ASuBOsbP1kgVDPVlQfbRw1"

# 1. Get Mishwa's contacts from Firestore
contacts_ref = db.collection("contacts")
mishwa_contacts = [doc.to_dict() | {"id": doc.id} for doc in contacts_ref.where("assignedTo", "==", mishwa_uid).get()]
print(f"Loaded {len(mishwa_contacts)} contacts assigned to Mishwa.")

# Create lookup sets
mishwa_emails = {c.get("email", "").strip().lower() for c in mishwa_contacts if c.get("email")}
mishwa_mobiles_last10 = set()
for c in mishwa_contacts:
    mob = re.sub(r'\D', '', c.get("mobile", ""))
    if len(mob) >= 10:
        mishwa_mobiles_last10.add(mob[-10:])

print(f"Mishwa Emails count: {len(mishwa_emails)}")
print(f"Mishwa Mobiles count: {len(mishwa_mobiles_last10)}")

# 2. Iterate through Excel files and check for matches
DOWNLOADS_DIR = "/Users/tirthpatel/Downloads"
excel_files = [f for f in os.listdir(DOWNLOADS_DIR) if f.endswith(".xlsx") and "recipient" in f.lower()]

matched_rows = []

for file in excel_files:
    file_path = os.path.join(DOWNLOADS_DIR, file)
    print(f"\nChecking Excel file: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    continue  # skip headers
                # Check each cell in the row
                for col_idx, val in enumerate(row, 1):
                    if not val:
                        continue
                    val_str = str(val).strip().lower()
                    
                    # Match email
                    is_email_match = val_str in mishwa_emails
                    
                    # Match phone
                    val_digits = re.sub(r'\D', '', val_str)
                    is_phone_match = len(val_digits) >= 10 and val_digits[-10:] in mishwa_mobiles_last10
                    
                    if is_email_match or is_phone_match:
                        matched_rows.append({
                            "file": file,
                            "sheet": sheet_name,
                            "row": row_idx,
                            "value": val,
                            "row_data": row
                        })
                        print(f"Match found in row {row_idx}: {row}")
    except Exception as e:
        print(f"Error reading Excel {file_path}: {e}")

print(f"\n=== Total Excel Matches: {len(matched_rows)} ===")
