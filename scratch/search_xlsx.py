import os
import openpyxl

DOWNLOADS_DIR = "/Users/tirthpatel/Downloads"

def search_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for col_idx, cell_value in enumerate(row, 1):
                    if cell_value:
                        val_str = str(cell_value).lower()
                        if "mishwa" in val_str or "hansaliya" in val_str or "marwadi" in val_str:
                            print(f"Match found in {file_path} | Sheet: {sheet_name} | Row {row_idx}, Col {col_idx} | Value: {cell_value}")
    except Exception as e:
        # If openpyxl is not installed or file is corrupt, print the error
        print(f"Error reading {file_path}: {e}")

def main():
    print("Searching Excel files in Downloads...")
    for file in os.listdir(DOWNLOADS_DIR):
        if file.endswith(".xlsx"):
            full_path = os.path.join(DOWNLOADS_DIR, file)
            search_excel(full_path)
    print("Excel scan complete.")

if __name__ == "__main__":
    main()
