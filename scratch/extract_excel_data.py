import os
import openpyxl
import re
import json

DOWNLOADS_DIR = "/Users/tirthpatel/Downloads"

def main():
    print("Extracting data from recipient certificate Excel files...")
    excel_files = [f for f in os.listdir(DOWNLOADS_DIR) if f.endswith(".xlsx") and "recipient" in f.lower()]
    
    extracted_emails = set()
    extracted_phones = set()
    
    for file in excel_files:
        file_path = os.path.join(DOWNLOADS_DIR, file)
        print(f"Reading {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for cell in row:
                        if not cell:
                            continue
                        cell_str = str(cell).strip()
                        
                        # Extract email pattern
                        email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', cell_str)
                        if email_match:
                            extracted_emails.add(email_match.group(0).lower())
                            
                        # Extract phone pattern (looking for groups of digits)
                        digits = re.sub(r'\D', '', cell_str)
                        if len(digits) >= 10:
                            extracted_phones.add(digits)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            
    result = {
        "emails": list(extracted_emails),
        "phones": list(extracted_phones)
    }
    
    output_path = "/Users/tirthpatel/Downloads/team-app-main/scratch/excel_extracted.json"
    with open(output_path, "w") as f:
        json.dump(result, f, indent=2)
        
    print(f"Extracted {len(extracted_emails)} unique emails and {len(extracted_phones)} unique phone numbers.")
    print(f"Saved to {output_path}")

if __name__ == "__main__":
    main()
